import argparse
import subprocess
import sys
import json
import os
try:
    import pandas as pd
except:
    subprocess.Popen([sys.executable, "-m", "pip", "install", "pandas"], creationflags=subprocess.CREATE_NEW_CONSOLE).wait()
    import pandas as pd

try:
    import openpyxl
except:
    subprocess.Popen([sys.executable, "-m", "pip", "install", "openpyxl"], creationflags=subprocess.CREATE_NEW_CONSOLE).wait()
    import openpyxl


class XlsxTable:
    def __init__(self, df):
        rows, cols = df.shape
        self.row = 0
        self.col = 0
        self.max_row = rows
        self.max_col = cols
        self.iloc = df.iloc
    def next(self):
        self.col += 1
        if self.col >= self.max_col:
            self.col = 0
            self.row += 1
        if self.row >= self.max_row:
            raise StopIteration
    def __str__(self):
        return f"({self.row}, {self.col})"
    
    def find_str_in_row(self, row, s):
        for c in range(self.max_col):
            if str(self.iloc[row, c]) == s:
                return c
        return -1
    
    def get_row(self, row):
        return self.iloc[row, :].tolist()
    
    def get_col(self, col):
        return self.iloc[:, col].tolist()
    
    def find_next_str_in_col(self,row, col, s):
        for r in range(row, self.max_row):
            if str(self.iloc[r, col]) == s:
                return r
        return -1
    
    def find_empty_in_col(self, row, col):
        for r in range(row, self.max_row):
            value = self.iloc[r, col]
            if value is None or pd.isna(value):
                return r
        return -1

# 读



if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=str, required=True, help="Path to xlsx")
    parser.add_argument("--json", type=str, required=True, help="Path to json")
    args = parser.parse_args()
    xlsx_file = args.xlsx
    if not os.path.exists(xlsx_file):
        xlsx_file = os.path.join(os.getcwd(), xlsx_file)
        if not os.path.exists(xlsx_file):
            print("文件不存在")
            sys.exit(1)
    df = pd.read_excel(xlsx_file, engine="openpyxl", header=None)
    result = []
    xlsx = XlsxTable(df)
    header_row = xlsx.get_row(4)
    header = {}
    for i, v in enumerate(header_row):
        if v is not None and not pd.isna(v) and not str(v).startswith("$"):
            header[v] = i
    data_type = xlsx.get_row(3)
    start = xlsx.find_next_str_in_col(0, 0,"Start")
    end = xlsx.find_next_str_in_col(0, 0,"End")
    output = {}
    for r in range(start, end):
        row_data = {}
        for k, c in header.items():
            if data_type[c] == "int":
                cell_value = xlsx.iloc[r, c]
                if pd.isna(cell_value):
                    row_data[k] = 0
                else:
                    row_data[k] = int(cell_value)
            elif data_type[c] == "string":
                cell_value = xlsx.iloc[r, c]
                if pd.isna(cell_value):
                    row_data[k] = ""
                else:
                    row_data[k] = str(cell_value)
            elif data_type[c] == "int[]":
                cell_value = xlsx.iloc[r, c]
                if pd.isna(cell_value):
                    row_data[k] = []
                else:
                    row_data[k] = [int(x) for x in str(cell_value)[1:-1].split(",")]
            elif data_type[c] == "string[]":
                cell_value = xlsx.iloc[r, c]
                if pd.isna(cell_value):
                    row_data[k] = []
                else:
                    row_data[k] = [str(x) for x in str(cell_value)[1:-1].split(",")]
            elif data_type[c] == "array":
                cell_value = xlsx.iloc[r, c]
                if pd.isna(cell_value):
                    row_data[k] = []
                else:
                    row_data[k] = json.loads(cell_value)
        output[row_data["Id"]] = row_data
    with open(args.json, 'w', encoding='utf-8') as f_out:
        json.dump(output, f_out, ensure_ascii=False, indent=2)
    print(f"处理表 : {args.xlsx}")