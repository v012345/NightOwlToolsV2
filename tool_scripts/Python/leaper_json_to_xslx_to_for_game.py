import argparse
import subprocess
import sys
import json
import os
try:
    import pandas as pd
except:
    subprocess.Popen([sys.executable, "-m", "pip", "install", "pandas"], creationflags=subprocess.CREATE_NEW_CONSOLE).wait()
    import pandas as pd

try:
    import openpyxl
except:
    subprocess.Popen([sys.executable, "-m", "pip", "install", "openpyxl"], creationflags=subprocess.CREATE_NEW_CONSOLE).wait()
    import openpyxl


class XlsxTable:
    def __init__(self, df):
        rows, cols = df.shape
        self.row = 0
        self.col = 0
        self.max_row = rows
        self.max_col = cols
        self.iloc = df.iloc
    def next(self):
        self.col += 1
        if self.col >= self.max_col:
            self.col = 0
            self.row += 1
        if self.row >= self.max_row:
            raise StopIteration
    def __str__(self):
        return f"({self.row}, {self.col})"
    
    def find_str_in_row(self, row, s):
        for c in range(self.max_col):
            if str(self.iloc[row, c]) == s:
                return c
        return -1
    
    def get_row(self, row):
        return self.iloc[row, :].tolist()
    
    def get_col(self, col):
        return self.iloc[:, col].tolist()
    
    def find_next_str_in_col(self,row, col, s):
        for r in range(row, self.max_row):
            if str(self.iloc[r, col]) == s:
                return r
        return -1
    
    def find_empty_in_col(self, row, col):
        for r in range(row, self.max_row):
            value = self.iloc[r, col]
            if value is None or pd.isna(value):
                return r
        return -1

def fix_float(v):
    """float 1.0, 2.0 转为 int"""
    if isinstance(v, float) and v == int(v):
        return int(v)
    if isinstance(v, list):
        return [fix_float(x) for x in v]
    if isinstance(v, dict):
        return {k: fix_float(val) for k, val in v.items()}
    return v


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=str, required=True, help="Path to xlsx")
    parser.add_argument("--json", type=str, required=True, help="Path to json")
    args = parser.parse_args()

    xlsx_file = args.xlsx
    if not os.path.exists(xlsx_file):
        xlsx_file = os.path.join(os.getcwd(), xlsx_file)
        if not os.path.exists(xlsx_file):
            print("xlsx 文件不存在")
            sys.exit(1)
    json_file = args.json
    if not os.path.exists(json_file):
        json_file = os.path.join(os.getcwd(), json_file)
        if not os.path.exists(json_file):
            print("json 文件不存在")
            sys.exit(1)

    with open(json_file, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    json_data = fix_float(json_data)

    # 用 pandas 读取结构信息
    df = pd.read_excel(xlsx_file, engine="openpyxl", header=None)
    xlsx = XlsxTable(df)
    header_row = xlsx.get_row(4)
    header = {}
    for i, v in enumerate(header_row):
        if v is not None and not pd.isna(v):
            header[v] = i
    data_type = xlsx.get_row(3)
    start = xlsx.find_next_str_in_col(0, 0, "Start")
    end = xlsx.find_next_str_in_col(0, 0, "End")

    # 用 openpyxl 打开写入，保留格式
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active

    # openpyxl 是 1-indexed，pandas 是 0-indexed
    row_offset = 1
    col_offset = 1

    json_items = list(json_data.values())
    write_row = start
    for item in json_items:
        if write_row >= end:
            break
        for k, c in header.items():

            if k not in item:
                continue
            value = item[k]
            cell = ws.cell(row=write_row + row_offset, column=c + col_offset)
            dt = data_type[c]
            if dt == "int":
                cell.value = int(value) if value else 0
            elif dt == "string":
                cell.value = str(value) if value else ""
            elif dt in ("int[]", "string[]"):
                if not value or len(value) == 0:
                    cell.value = None
                else:
                    cell.value = "[" + ",".join(str(x) for x in value) + "]"
            elif dt == "array":
                if not value or len(value) == 0:
                    cell.value = None
                else:
                    cell.value = json.dumps(value, ensure_ascii=False).replace(" ", "")
            else:
                cell.value = value
        write_row += 1

    wb.save(xlsx_file)
    print(f"写回表 : {args.xlsx}")